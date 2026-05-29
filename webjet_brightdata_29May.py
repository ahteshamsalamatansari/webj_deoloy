"""
Webjet.com.au Flight Scraper — Bright Data Browser API
=======================================================
Routes  : PER → MJK  (Perth → Monkey Mia)
          MJK → PER  (Monkey Mia → Perth)

Method  : Navigate to Webjet flight matrix URL (one per date/route).
          Bright Data Scraping Browser handles Cloudflare automatically.
          Extract: ZL flight number, departure time, fare price, fare class.

Output  : webjet_results.xlsx  (resume-friendly, atomic writes)
Cron    : Render.com cron job — no interactive input required.
          Same Bright Data credentials as rex_brightdata_Orignal1_updated.py.
"""

import asyncio
import asyncio
import os
import re
import sys
import time
import argparse
import json
import random
import tempfile
import traceback
import urllib.request
from dataclasses import dataclass, field as dc_field
from datetime import datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError
from playwright.async_api import async_playwright, Browser
from openpyxl import load_workbook, Workbook

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

ORIGINAL_STDOUT = sys.stdout
ORIGINAL_STDERR = sys.stderr

# ─────────────────────────────────────────────────────────────
#  BRIGHT DATA CREDENTIALS  (same zone as Rex scraper)
# ─────────────────────────────────────────────────────────────

BD_BROWSER_HOST = os.getenv("BD_BROWSER_HOST", "brd.superproxy.io")
BD_BROWSER_PORT = os.getenv("BD_BROWSER_PORT", "9222")
BD_BROWSER_USER = os.getenv(
    "BD_BROWSER_USER",
    "brd-customer-hl_fbc4a16a-zone-cont_rex",
)
BD_BROWSER_PASS = os.getenv("BD_BROWSER_PASS", "072res2p22t3")
BD_BROWSER_WSS  = os.getenv(
    "BD_BROWSER_WSS",
    f"wss://{BD_BROWSER_USER}:{BD_BROWSER_PASS}@{BD_BROWSER_HOST}:{BD_BROWSER_PORT}",
)

SENSITIVE_BD_RE = re.compile(
    r"(?:wss|https)://[^\s'\"<>]+@brd\.superproxy\.io:\d+/?"
)


def redact_sensitive_text(text: str) -> str:
    return SENSITIVE_BD_RE.sub(
        "brightdata://<redacted>@brd.superproxy.io", str(text)
    )


# ─────────────────────────────────────────────────────────────
#  CAPSOLVER — Cloudflare Turnstile / reCAPTCHA fallback solver
#  Used when Bright Data's built-in solver returns "failed" for
#  Cloudflare challenges on the Webjet page.
# ─────────────────────────────────────────────────────────────
CAPSOLVER_API_KEY = os.getenv(
    "CAPSOLVER_API_KEY",
    "CAP-D9EC80AA2D25FEB038515477B6B0F668AFE90492C3A19ECA7EFF0D31969EE410",
)


def _capsolver_post_sync(url: str, payload: dict) -> dict:
    """Synchronous HTTP POST to CapSolver REST API (called via asyncio.to_thread)."""
    data = json.dumps(payload).encode()
    req = urllib.request.Request(
        url, data=data, headers={"Content-Type": "application/json"}
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read().decode())


async def capsolver_solve_recaptcha_v2(
    site_key: str, page_url: str, timeout: int = 90
) -> "str | None":
    """
    Solve a Google reCAPTCHA v2 via CapSolver API.
    Returns the gRecaptchaResponse token, or None on failure.
    """
    if not CAPSOLVER_API_KEY:
        return None
    try:
        create_resp = await asyncio.to_thread(
            _capsolver_post_sync,
            "https://api.capsolver.com/createTask",
            {
                "clientKey": CAPSOLVER_API_KEY,
                "task": {
                    "type": "ReCaptchaV2TaskProxyless",
                    "websiteURL": page_url,
                    "websiteKey": site_key,
                },
            },
        )
        if create_resp.get("errorId", 0) != 0:
            print(f"   ❌ CapSolver reCAPTCHA error: {create_resp.get('errorDescription')}")
            return None
        task_id = create_resp.get("taskId")
        if not task_id:
            return None
        print(f"   🔑 CapSolver reCAPTCHA task queued (id={task_id[:8]}...)")

        loop = asyncio.get_running_loop()
        deadline = loop.time() + timeout
        while loop.time() < deadline:
            await asyncio.sleep(4)
            result_resp = await asyncio.to_thread(
                _capsolver_post_sync,
                "https://api.capsolver.com/getTaskResult",
                {"clientKey": CAPSOLVER_API_KEY, "taskId": task_id},
            )
            if result_resp.get("errorId", 0) != 0:
                print(f"   ❌ CapSolver poll error: {result_resp.get('errorDescription')}")
                return None
            if result_resp.get("status") == "ready":
                token = result_resp.get("solution", {}).get("gRecaptchaResponse")
                if token:
                    print("   ✅ CapSolver: reCAPTCHA v2 solved!")
                    return token
                return None
        print("   ⏱️  CapSolver: reCAPTCHA timed out")
        return None
    except Exception as exc:
        print(f"   ❌ CapSolver reCAPTCHA exception: {exc}")
        return None


async def capsolver_solve_turnstile(
    site_key: str, page_url: str, timeout: int = 90
) -> "str | None":
    """
    Solve a Cloudflare Turnstile challenge via CapSolver API.
    Returns the Turnstile token, or None on failure.
    """
    if not CAPSOLVER_API_KEY:
        return None
    try:
        create_resp = await asyncio.to_thread(
            _capsolver_post_sync,
            "https://api.capsolver.com/createTask",
            {
                "clientKey": CAPSOLVER_API_KEY,
                "task": {
                    "type": "AntiTurnstileTaskProxyLess",
                    "websiteURL": page_url,
                    "websiteKey": site_key,
                },
            },
        )
        if create_resp.get("errorId", 0) != 0:
            print(f"   ❌ CapSolver Turnstile error: {create_resp.get('errorDescription')}")
            return None
        task_id = create_resp.get("taskId")
        if not task_id:
            return None
        print(f"   🔑 CapSolver Turnstile task queued (id={task_id[:8]}...)")

        loop = asyncio.get_running_loop()
        deadline = loop.time() + timeout
        while loop.time() < deadline:
            await asyncio.sleep(4)
            result_resp = await asyncio.to_thread(
                _capsolver_post_sync,
                "https://api.capsolver.com/getTaskResult",
                {"clientKey": CAPSOLVER_API_KEY, "taskId": task_id},
            )
            if result_resp.get("errorId", 0) != 0:
                print(f"   ❌ CapSolver Turnstile poll error: {result_resp.get('errorDescription')}")
                return None
            if result_resp.get("status") == "ready":
                token = result_resp.get("solution", {}).get("token")
                if token:
                    print("   ✅ CapSolver: Cloudflare Turnstile solved!")
                    return token
                return None
        print("   ⏱️  CapSolver: Turnstile timed out")
        return None
    except Exception as exc:
        print(f"   ❌ CapSolver Turnstile exception: {exc}")
        return None


# ─────────────────────────────────────────────────────────────
#  CONFIGURATION
# ─────────────────────────────────────────────────────────────

AIRPORT_MAP = {
    "PER": "Perth",
    "MJK": "Monkey Mia",
}

WEBJET_ROUTES = [
    ("PER", "MJK"),
    ("MJK", "PER"),
]

# Route-specific URL components
ROUTE_CONFIG: dict[tuple[str, str], dict] = {
    ("PER", "MJK"): {
        "CityFrom":   "Perth",
        "CityTo":     "Monkey Mia",
        "OneWayFmt":  "PER-PER-MJK-MJK-{date}",
    },
    ("MJK", "PER"): {
        "CityFrom":   "Monkey Mia",
        "CityTo":     "Perth",
        "OneWayFmt":  "MJK-MJK-PER-PER-{date}",
    },
}

# Timezone (same as Rex scraper)
_WJ_TZ_NAME = os.getenv("REX_TIMEZONE", "Australia/Perth")
try:
    WJ_TZ = ZoneInfo(_WJ_TZ_NAME)
except ZoneInfoNotFoundError:
    _WJ_TZ_NAME = "Australia/Perth"
    WJ_TZ = ZoneInfo(_WJ_TZ_NAME)

TOTAL_DAYS              = int(os.getenv("WJ_TOTAL_DAYS", "84"))
START_OFFSET_DAYS       = int(os.getenv("WJ_START_OFFSET_DAYS", "1"))
OUTPUT_EXCEL            = os.getenv("WJ_OUTPUT_EXCEL", "webjet_results.xlsx")
DEBUG_DIR               = os.getenv("WJ_DEBUG_DIR", "webjet_debug")
LOG_DIR                 = os.getenv("WJ_LOG_DIR", "webjet_logs")
RUN_ID                  = os.getenv("WJ_RUN_ID",
                              datetime.now(WJ_TZ).strftime("%Y%m%d"))

MAX_ATTEMPTS            = int(os.getenv("WJ_MAX_ATTEMPTS", "3"))
FINAL_RETRY_ROUNDS      = int(os.getenv("WJ_FINAL_RETRY_ROUNDS", "1"))
RETRY_BACKOFF_SECONDS   = float(os.getenv("WJ_RETRY_BACKOFF_SECONDS", "8"))
MAX_RETRY_BACKOFF_SECS  = float(os.getenv("WJ_MAX_RETRY_BACKOFF_SECONDS", "90"))
JOB_TIMEOUT_SECONDS     = int(os.getenv("WJ_JOB_TIMEOUT_SECONDS", "240"))
RESUME_ENABLED          = os.getenv("WJ_RESUME", "1").lower() not in {"0", "false", "no"}
INTER_ROUTE_DELAY_SECS  = int(os.getenv("WJ_INTER_ROUTE_DELAY_SECONDS", "30"))
CHECKPOINT_EVERY        = int(os.getenv("WJ_CHECKPOINT_EVERY", "7"))
PAGE_LOAD_TIMEOUT       = int(os.getenv("WJ_PAGE_LOAD_TIMEOUT", "90"))

CSV_FIELDS = [
    "Date Checked", "Time Checked", "Airline",
    "Date of Departure", "Time of Departure",
    "Origin", "Destination",
    "Fare Price", "Fare Class", "Source",
]

# Internal tracking column written to Excel but NOT in the user-visible list.
# "Status" is required for resume / cleanup-pass logic (job_completed and
# failed_route_dates both read it back from Excel rows).
# "Run ID" is intentionally excluded — row matching now uses Origin+Dest+Date only.
INTERNAL_EXCEL_FIELDS = ["Status"]

# Status constants (mirror Rex scraper for consistent Excel output)
STATUS_SUCCESS          = "SUCCESS"
STATUS_NO_FARE          = "NO_FARE_AVAILABLE"
STATUS_FAILED           = "FAILED_AFTER_RETRIES"
STATUS_SITE_UNAVAILABLE = "SITE_UNAVAILABLE"
STATUS_STRUCTURE        = "POSSIBLE_WEBSITE_STRUCTURE_ISSUE"
STATUS_TIMEOUT          = "PAGE_TIMEOUT"
STATUS_BLOCKED          = "BLOCKED_OR_VERIFICATION_REQUIRED"
STATUS_ROUTE_SETUP_FAILED = "ROUTE_SETUP_FAILED"

COMPLETED_STATUSES = {STATUS_SUCCESS, STATUS_NO_FARE}
RETRYABLE_FAILURE_STATUSES = {
    STATUS_FAILED,
    STATUS_SITE_UNAVAILABLE,
    STATUS_STRUCTURE,
    STATUS_TIMEOUT,
    STATUS_BLOCKED,
    STATUS_ROUTE_SETUP_FAILED,
}

# Webjet matrix table selectors (based on actual page HTML structure)
# Primary:  table.matrix-table > tbody > tr.result-row
# Fallback: tr.result-row (without full table context)
MATRIX_TABLE_SEL    = "table.matrix-table"
RESULT_ROW_SEL      = "tr.result-row"
FARE_COL_SEL        = "thead th.header-col"
FARE_TYPE_TEXT_SEL  = "span.fare-type-text"
TIME_DEPART_SEL     = "td.left-header span.time-depart span"
TIME_ARRIVE_SEL     = "td.left-header span.time-arrive span"
FARE_PRICE_CELL_SEL = "td.fare-price"
PRICE_TEXT_SEL      = "span.price-text"


# ─────────────────────────────────────────────────────────────
#  DATE / URL UTILITIES
# ─────────────────────────────────────────────────────────────

def wj_now() -> datetime:
    """Current date/time in WJ timezone (naive)."""
    return datetime.now(WJ_TZ).replace(tzinfo=None)


def today_dt() -> datetime:
    return wj_now().replace(hour=0, minute=0, second=0, microsecond=0)


def build_date_list() -> list[datetime]:
    t = today_dt() + timedelta(days=START_OFFSET_DAYS)
    return [t + timedelta(days=i) for i in range(TOTAL_DAYS)]


def output_date(dt: datetime) -> str:
    return dt.strftime("%d-%m-%Y")


def build_webjet_url(origin: str, dest: str, dt: datetime) -> str:
    """Build the Webjet flight matrix URL for a specific route and date."""
    cfg      = ROUTE_CONFIG[(origin, dest)]
    date_str = dt.strftime("%Y%m%d")
    one_way  = cfg["OneWayFmt"].format(date=date_str)
    city_from = cfg["CityFrom"].replace(" ", "%20")
    city_to   = cfg["CityTo"].replace(" ", "%20")
    return (
        "https://services.webjet.com.au/web/flights/matrix/"
        f"?Adults=1&Children=0&Infants=0&TravelClass=Economy"
        f"&GeoCategory=AUDomestic&TripType=Oneway"
        f"&OneWay={one_way}"
        f"&CityFrom={city_from}&CityTo={city_to}"
        f"&CityCodeFrom={origin}&CityCodeTo={dest}"
    )


# ─────────────────────────────────────────────────────────────
#  PRICE / TIME / FARE CLASS EXTRACTION
# ─────────────────────────────────────────────────────────────

def _ensure_cents(val: str) -> str:
    return val if "." in val else f"{val}.00"


def extract_price_from_text(text: str) -> str:
    """Extract lowest/first price from a text segment."""
    # "From $196.00" or "from $196.00"
    m = re.search(r"[Ff]rom\s*\$\s*([\d,]+\.\d{2})", text)
    if m:
        return f"${m.group(1)}"
    # "$196.00"
    m = re.search(r"\$\s*([\d,]+\.\d{2})", text)
    if m:
        return f"${m.group(1)}"
    # "$196" (no cents)
    m = re.search(r"\$\s*(\d[\d,]+)", text)
    if m:
        return f"${_ensure_cents(m.group(1))}"
    return "N/A"


def extract_time_from_text(text: str) -> str:
    """Extract the first time (12-hr or 24-hr) from text."""
    # 12-hour with am/pm:  9:50am  10:30 AM  10:30am
    m = re.search(r"\b(\d{1,2}:\d{2}\s*[aApP][mM])\b", text)
    if m:
        return m.group(1).strip()
    # 24-hour:  09:50  14:35
    m = re.search(r"\b(\d{1,2}:\d{2})\b", text)
    if m:
        return m.group(1)
    return "-"


def extract_fare_class_from_text(text: str) -> str:
    """Extract fare class label from text, defaulting to 'Economy'."""
    patterns = [
        r"(Economy\s+(?:Saver|Flex|Flexi|Value|Plus|Standard|Light|Basic|Classic|Full))",
        r"(Business\s+(?:Saver|Flex|Flexi|Value|Standard|Class)?)",
        r"\b(Saver|Flexi|Flex|Value|Standard|Plus|Light|Basic|Premium|Classic)\b",
    ]
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            return m.group(1).strip().title()
    return "Economy"


# ─────────────────────────────────────────────────────────────
#  LOGGING (TeeStream)
# ─────────────────────────────────────────────────────────────

class TeeStream:
    """Duplicate stdout/stderr to a log file while keeping console output."""

    def __init__(self, *streams):
        self.streams = streams

    def write(self, text):
        safe = redact_sensitive_text(text)
        for s in self.streams:
            try:
                s.write(safe)
                s.flush()
            except (ValueError, IOError, OSError):
                pass

    def flush(self):
        for s in self.streams:
            try:
                s.flush()
            except (ValueError, IOError, OSError):
                pass

    def reconfigure(self, **kwargs):
        for s in self.streams:
            if hasattr(s, "reconfigure"):
                s.reconfigure(**kwargs)


def configure_run_logging(log_dir: str, run_id: str):
    if not log_dir:
        return None
    os.makedirs(log_dir, exist_ok=True)
    log_path = os.path.join(
        log_dir,
        f"webjet_{run_id}_{wj_now():%Y%m%d_%H%M%S}.log"
    )
    log_fh = open(log_path, "a", encoding="utf-8", errors="replace")
    sys.stdout = TeeStream(sys.stdout, log_fh)
    sys.stderr = TeeStream(sys.stderr, log_fh)
    print(f"🧾 Cron log: {os.path.abspath(log_path)}")
    return log_fh


def restore_run_logging(log_fh):
    if not log_fh:
        return
    try:
        sys.stdout.flush()
        sys.stderr.flush()
    except Exception:
        pass
    sys.stdout = ORIGINAL_STDOUT
    sys.stderr = ORIGINAL_STDERR
    try:
        log_fh.flush()
        log_fh.close()
    except Exception:
        pass


# ─────────────────────────────────────────────────────────────
#  OUTPUT STORE  (resume-friendly, atomic Excel writer)
# ─────────────────────────────────────────────────────────────

class OutputStore:
    """Atomic, resume-friendly Excel writer keyed by run/origin/dest/date."""

    def __init__(self, path: str, run_id: str):
        self.path   = path
        self.run_id = run_id
        self._entries_since_checkpoint = 0

    def _load(self):
        if os.path.exists(self.path):
            wb = load_workbook(self.path)
            ws = wb.active
            first_row_has_values = any(
                ws.cell(row=1, column=c).value
                for c in range(1, len(CSV_FIELDS) + 1)
            )
            if not first_row_has_values:
                ws.append(CSV_FIELDS)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Webjet Flight Data"
            ws.append(CSV_FIELDS)
        self._ensure_headers(ws)
        return wb, ws

    def _headers(self, ws) -> list[str]:
        return [cell.value or "" for cell in ws[1]]

    def _ensure_headers(self, ws):
        headers = self._headers(ws)
        if not any(headers):
            # Brand-new or fully-empty header row — write all columns from scratch.
            for idx, col in enumerate(CSV_FIELDS, 1):
                ws.cell(row=1, column=idx).value = col
            # Internal tracking columns go directly after the user-visible ones.
            for offset, col in enumerate(INTERNAL_EXCEL_FIELDS, 1):
                ws.cell(row=1, column=len(CSV_FIELDS) + offset).value = col
            return
        changed = False
        for col in CSV_FIELDS + INTERNAL_EXCEL_FIELDS:
            if col not in headers:
                headers.append(col)
                ws.cell(row=1, column=len(headers)).value = col
                changed = True
        if changed:
            print("   ℹ️  Added new columns to existing workbook.")

    def _save_atomic(self, wb):
        out_dir = os.path.dirname(os.path.abspath(self.path)) or "."
        os.makedirs(out_dir, exist_ok=True)
        fd, tmp = tempfile.mkstemp(
            prefix=f".{Path(self.path).stem}_",
            suffix=".xlsx",
            dir=out_dir,
        )
        os.close(fd)
        try:
            wb.save(tmp)
            os.replace(tmp, self.path)
        finally:
            if os.path.exists(tmp):
                try:
                    os.remove(tmp)
                except OSError:
                    pass

    def _row_matches(self, ws, row_idx, run_id, origin, dest, date_str) -> bool:
        # NOTE: run_id parameter kept for signature compatibility but is no longer
        # used for matching — "Run ID" was removed from the user-visible output.
        # Rows are uniquely identified by Origin + Destination + Date of Departure.
        headers = self._headers(ws)
        vals = {
            h: ws.cell(row=row_idx, column=i + 1).value
            for i, h in enumerate(headers)
        }
        return (
            str(vals.get("Origin") or "") == origin
            and str(vals.get("Destination") or "") == dest
            and str(vals.get("Date of Departure") or "") == date_str
        )

    def write_job_rows(self, origin: str, dest: str, date_str: str,
                       rows: list[dict]):
        wb, ws = self._load()
        headers = self._headers(ws)

        for r in range(ws.max_row, 1, -1):
            if self._row_matches(ws, r, self.run_id, origin, dest, date_str):
                ws.delete_rows(r, 1)

        for row in rows:
            row.setdefault("Run ID", self.run_id)
            row.setdefault("Status", STATUS_SUCCESS)
            row.setdefault("Comment", "")
            row.setdefault("Retry Count", 0)
            row.setdefault("Debug Artifacts", "")
            ws.append([row.get(f, "") for f in headers])
            self._entries_since_checkpoint += 1

        if self._entries_since_checkpoint >= CHECKPOINT_EVERY:
            self._save_atomic(wb)
            print(
                f"   💾 Checkpoint: {self._entries_since_checkpoint} "
                f"entries saved → {self.path}"
            )
            self._entries_since_checkpoint = 0
        else:
            self._save_atomic(wb)

    def job_rows(self, origin: str, dest: str, date_str: str) -> list[dict]:
        if not os.path.exists(self.path):
            return []
        wb, ws = self._load()
        headers = self._headers(ws)
        rows = []
        for r in range(2, ws.max_row + 1):
            if self._row_matches(ws, r, self.run_id, origin, dest, date_str):
                rows.append({
                    h: ws.cell(row=r, column=i + 1).value
                    for i, h in enumerate(headers)
                })
        return rows

    def job_completed(self, origin: str, dest: str, date_str: str) -> bool:
        rows = self.job_rows(origin, dest, date_str)
        if not rows:
            return False
        return all((r.get("Status") or "") in COMPLETED_STATUSES for r in rows)

    def job_has_any_row(self, origin: str, dest: str, date_str: str) -> bool:
        return bool(self.job_rows(origin, dest, date_str))

    def failed_route_dates(self, origin: str, dest: str,
                           date_strs: list[str]) -> list[str]:
        """Return date_strs that have retryable failures (not completed)."""
        failed = []
        for ds in date_strs:
            if self.job_completed(origin, dest, ds):
                continue
            rows = self.job_rows(origin, dest, ds)
            if not rows:
                continue
            statuses = {(r.get("Status") or "") for r in rows}
            if statuses & RETRYABLE_FAILURE_STATUSES:
                failed.append(ds)
        return failed


OUTPUT_STORE: OutputStore | None = None


def _append_rows(rows: list[dict]):
    """Fallback writer when OUTPUT_STORE is not initialised."""
    if not rows:
        return
    if os.path.exists(OUTPUT_EXCEL):
        wb = load_workbook(OUTPUT_EXCEL)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Webjet Flight Data"
        ws.append(CSV_FIELDS)
    for row in rows:
        ws.append([row.get(f, "") for f in CSV_FIELDS])
    wb.save(OUTPUT_EXCEL)


# ─────────────────────────────────────────────────────────────
#  DATA CLASS
# ─────────────────────────────────────────────────────────────

@dataclass
class JobResult:
    status: str
    rows: list[dict] = dc_field(default_factory=list)
    comment: str = ""
    retryable: bool = True
    retry_count: int = 0
    debug_artifacts: list[str] = dc_field(default_factory=list)
    # True when the no-flights signal is ambiguous (URL redirect or hiccup error)
    # and a fresh-IP verification attempt is worthwhile.
    needs_ip_verification: bool = False

    @property
    def completed(self) -> bool:
        return self.status in COMPLETED_STATUSES


# ─────────────────────────────────────────────────────────────
#  SCRAPER
# ─────────────────────────────────────────────────────────────

class WebjetScraper:

    # Markers that indicate the Bright Data WebSocket session has been dropped
    _BROWSER_CLOSED_MARKERS = (
        "target page, context or browser has been closed",
        "browser has been closed",
        "connection closed",
        "websocket",
        "target closed",
    )

    def __init__(
        self,
        debug_dir: str = DEBUG_DIR,
        max_attempts: int = MAX_ATTEMPTS,
        final_retry_rounds: int = FINAL_RETRY_ROUNDS,
        retry_backoff: float = RETRY_BACKOFF_SECONDS,
        max_retry_backoff: float = MAX_RETRY_BACKOFF_SECS,
        job_timeout: int = JOB_TIMEOUT_SECONDS,
        resume: bool = RESUME_ENABLED,
    ):
        self.debug_dir         = debug_dir
        self.max_attempts      = max(1, max_attempts)
        self.final_retry_rounds = max(0, final_retry_rounds)
        self.retry_backoff     = max(0.0, retry_backoff)
        self.max_retry_backoff = max(1.0, max_retry_backoff)
        self.job_timeout       = max(30, job_timeout)
        self.resume            = resume
        self._pw_instance      = None  # stored for _reconnect_browser()
        self._ip_clean         = False # True once current IP has loaded ≥1 SUCCESS

    # ── Browser lifecycle ────────────────────────────────────

    def _is_browser_closed_error(self, exc: Exception) -> bool:
        msg = str(exc).lower()
        return any(m in msg for m in self._BROWSER_CLOSED_MARKERS)

    async def _reconnect_browser(self) -> Browser | None:
        """Re-establish Bright Data CDP connection (fresh IP from pool)."""
        if not self._pw_instance:
            print("   ❌ Cannot reconnect — playwright instance not stored")
            return None
        print("   🔌 Reconnecting to Bright Data for a fresh IP...")
        for attempt in range(1, 4):
            try:
                new_browser = await self._pw_instance.chromium.connect_over_cdp(
                    BD_BROWSER_WSS
                )
                print(f"   ✅ Bright Data reconnected (attempt {attempt}/3)")
                self._ip_clean = False  # new IP — trust not yet established
                return new_browser
            except Exception as exc:
                print(f"   ⚠️  Reconnect {attempt}/3 failed: {exc}")
                if attempt < 3:
                    await asyncio.sleep(8 * attempt)
        print("   ❌ All reconnect attempts exhausted")
        return None

    async def _reconnect_with_new_page(
        self,
        eff: list,
        eff_ctx: list,
        eff_page: list,
        label: str = "reconnect",
    ) -> bool:
        """
        Reconnect Bright Data for a fresh IP, close the old context, and open
        a new context + page on the new browser.  Updates all three mutable
        refs in-place.  Returns True on success, False if reconnect fails.
        """
        # Close old context (also destroys its page)
        if eff_ctx[0]:
            await self._close_context(eff_ctx[0], f"{label} — old ctx")
            eff_ctx[0] = None
        eff_page[0] = None

        new_b = await self._reconnect_browser()
        if not new_b:
            return False

        eff[0] = new_b
        new_ctx, new_page = await self._new_context(new_b)
        eff_ctx[0] = new_ctx
        eff_page[0] = new_page
        return True

    async def _new_context(self, browser):
        """Create a browser context + page, reconnecting if the session died."""
        effective = browser
        for reconnect_no in range(3):
            try:
                ctx = await effective.new_context(
                    viewport={"width": 1366, "height": 900},
                    timezone_id=_WJ_TZ_NAME,
                    locale="en-AU",
                )
                ctx.set_default_timeout(30000)
                ctx.set_default_navigation_timeout(70000)
                page = await ctx.new_page()
                return ctx, page
            except Exception as exc:
                if reconnect_no < 2 and self._is_browser_closed_error(exc):
                    new_b = await self._reconnect_browser()
                    if new_b:
                        effective = new_b
                        continue
                raise

    async def _close_context(self, ctx, label: str = "context"):
        if ctx:
            try:
                await ctx.close()
            except Exception as exc:
                print(f"   ⚠️  {label} close error: {exc}")

    # ── Debug artifacts ──────────────────────────────────────

    async def save_debug(self, page, label: str,
                         metadata: dict | None = None) -> list[str]:
        os.makedirs(self.debug_dir, exist_ok=True)
        safe = re.sub(r"[^A-Za-z0-9_.-]+", "_", label).strip("_") or "page"
        ts   = wj_now().strftime("%Y%m%d_%H%M%S")
        artifacts: list[str] = []

        png_path = os.path.join(self.debug_dir, f"{safe}_{ts}.png")
        try:
            await page.screenshot(path=png_path, full_page=True)
            artifacts.append(os.path.abspath(png_path))
            print(f"   📸 Debug screenshot: {os.path.abspath(png_path)}")
        except Exception as exc:
            print(f"   ⚠️  Screenshot failed: {exc}")

        html_path = os.path.join(self.debug_dir, f"{safe}_{ts}.html")
        try:
            html = await page.content()
            with open(html_path, "w", encoding="utf-8") as fh:
                fh.write(html)
            artifacts.append(os.path.abspath(html_path))
            print(f"   🧾 Debug HTML: {os.path.abspath(html_path)}")
        except Exception as exc:
            print(f"   ⚠️  HTML dump failed: {exc}")

        meta = dict(metadata or {})
        try:
            meta.setdefault("url", page.url)
        except Exception:
            pass
        meta.setdefault("saved_at", wj_now().isoformat(sep=" "))
        try:
            meta["body_snippet"] = (
                await page.inner_text("body", timeout=3000)
            ).strip()[:1500]
        except Exception:
            pass

        json_path = os.path.join(self.debug_dir, f"{safe}_{ts}.json")
        try:
            with open(json_path, "w", encoding="utf-8") as fh:
                json.dump(meta, fh, ensure_ascii=False, indent=2)
            artifacts.append(os.path.abspath(json_path))
        except Exception as exc:
            print(f"   ⚠️  Debug JSON failed: {exc}")

        return artifacts

    # ── Cloudflare / page-state detection ───────────────────

    async def wait_for_brightdata_captcha(self, page,
                                          detect_timeout: int = 60000) -> str:
        """
        Ask Bright Data's Scraping Browser to solve any challenge.
        Returns "solved", "failed", or "not_solved".
        """
        try:
            client = await page.context.new_cdp_session(page)
            result = await client.send(
                "Captcha.waitForSolve", {"detectTimeout": detect_timeout}
            )
            status = (
                (result or {}).get("status", "unknown")
                if isinstance(result, dict)
                else "unknown"
            )
            if status in ("solved", "solve_finished"):
                print(f"   ✅ Bright Data solved challenge (status={status!r})")
                return "solved"
            elif status in ("solve_failed", "invalid"):
                print(f"   ❌ Challenge hard-failed (status={status!r})")
                return "failed"
            else:
                print(f"   ℹ️  Challenge status: {status!r}")
                return "not_solved"
        except Exception as exc:
            print(f"   ⚠️  Captcha wait exception: {exc}")
            return "not_solved"

    # ── CapSolver integration ────────────────────────────────

    async def _get_turnstile_sitekey(self, page) -> "str | None":
        """Extract the Cloudflare Turnstile data-sitekey from the page."""
        try:
            sitekey = await page.evaluate("""() => {
                // Turnstile widget container (div.cf-turnstile or any element with data-sitekey)
                const el = document.querySelector('.cf-turnstile[data-sitekey], [data-sitekey]');
                if (el) return el.getAttribute('data-sitekey');
                // Scan inline scripts for embedded sitekey
                for (const s of document.querySelectorAll('script')) {
                    const m = s.textContent.match(/['"](0x[0-9A-Fa-f]{15,})['"]/);
                    if (m) return m[1];
                }
                return null;
            }""")
            return sitekey or None
        except Exception:
            return None

    async def _get_recaptcha_sitekey(self, page) -> "str | None":
        """Extract a Google reCAPTCHA v2 site key from the page (fallback)."""
        try:
            sitekey = await page.evaluate("""() => {
                const el = document.querySelector('.g-recaptcha[data-sitekey], [data-sitekey]');
                if (el) {
                    const k = el.getAttribute('data-sitekey');
                    // reCAPTCHA site keys start with 6L (Turnstile keys start with 0x)
                    if (k && k.startsWith('6L')) return k;
                }
                return null;
            }""")
            return sitekey or None
        except Exception:
            return None

    async def _inject_turnstile_token(self, page, token: str) -> bool:
        """
        Inject a Cloudflare Turnstile token into the challenge page and submit the form.
        This replicates what the Turnstile widget would do after a real verification.
        """
        try:
            result = await page.evaluate("""(token) => {
                try {
                    // Set all known Turnstile response fields
                    ['cf-turnstile-response', '__cf_chl_f_tk', 'cf_captcha_kind'].forEach(name => {
                        document.querySelectorAll('[name="' + name + '"]').forEach(el => {
                            el.value = token;
                        });
                    });
                    // Also set the generic g-recaptcha-response if present (some CF pages use it)
                    const gr = document.getElementById('g-recaptcha-response');
                    if (gr) gr.value = token;
                    // Submit the challenge form
                    const form = document.getElementById('challenge-form')
                               || document.querySelector('form[action*="cdn-cgi"]')
                               || document.querySelector('form');
                    if (form) {
                        form.submit();
                        return 'form_submitted';
                    }
                    return 'no_form_found';
                } catch(e) {
                    return 'error:' + e.message;
                }
            }""", token)
            print(f"   🔑 Turnstile token injected (result={result!r})")
            await asyncio.sleep(3)
            return True
        except Exception as exc:
            print(f"   ⚠️  Turnstile token injection failed: {exc}")
            return False

    async def _inject_recaptcha_token_webjet(self, page, token: str) -> bool:
        """Inject a reCAPTCHA v2 token into the Webjet page and submit the form."""
        try:
            result = await page.evaluate("""(token) => {
                try {
                    const textarea = document.getElementById('g-recaptcha-response');
                    if (textarea) {
                        textarea.style.display = '';
                        textarea.value = token;
                    }
                    const form = document.querySelector('form');
                    if (form) { form.submit(); return 'form_submitted'; }
                    return 'no_form';
                } catch(e) {
                    return 'error:' + e.message;
                }
            }""", token)
            print(f"   🔑 reCAPTCHA token injected (result={result!r})")
            await asyncio.sleep(3)
            return True
        except Exception as exc:
            print(f"   ⚠️  reCAPTCHA token injection failed: {exc}")
            return False

    async def _try_capsolver_webjet(self, page, remaining_seconds: float) -> bool:
        """
        Attempt to solve Webjet's Cloudflare challenge via CapSolver.

        Strategy:
          1. Try Cloudflare Turnstile (AntiTurnstileTaskProxyLess) — the most common
             Webjet challenge type.
          2. Fall back to reCAPTCHA v2 (ReCaptchaV2TaskProxyless) if no Turnstile
             widget is found.

        Returns True if a token was obtained and injected (page should proceed).
        Returns False if CapSolver is unavailable or all solver attempts fail.
        """
        if not CAPSOLVER_API_KEY:
            return False

        page_url = page.url
        # Leave 5 s after CapSolver returns for token injection + page load
        capsolver_budget = max(20, min(90, int(remaining_seconds) - 5))

        # ── Attempt 1: Cloudflare Turnstile ─────────────────
        turnstile_key = await self._get_turnstile_sitekey(page)
        if turnstile_key:
            print(
                f"   🔄 CapSolver: Turnstile challenge detected "
                f"(sitekey={turnstile_key[:12]}..., budget={capsolver_budget}s)"
            )
            token = await capsolver_solve_turnstile(turnstile_key, page_url, timeout=capsolver_budget)
            if token:
                return await self._inject_turnstile_token(page, token)
            print("   ⚠️  CapSolver Turnstile failed — trying reCAPTCHA fallback...")

        # ── Attempt 2: reCAPTCHA v2 ─────────────────────────
        recaptcha_key = await self._get_recaptcha_sitekey(page)
        if recaptcha_key:
            print(
                f"   🔄 CapSolver: reCAPTCHA v2 detected "
                f"(sitekey={recaptcha_key[:12]}..., budget={capsolver_budget}s)"
            )
            token = await capsolver_solve_recaptcha_v2(recaptcha_key, page_url, timeout=capsolver_budget)
            if token:
                return await self._inject_recaptcha_token_webjet(page, token)

        print("   ⚠️  CapSolver: no recognisable challenge widget found on Webjet page")
        return False

    async def _page_is_cloudflare(self, page) -> bool:
        try:
            body  = (await page.inner_text("body", timeout=3000)).lower()
            title = (await page.title()).lower()
            markers = [
                "cloudflare", "verify you are human",
                "checking your browser", "just a moment",
                "turnstile", "challenge-platform", "captcha",
            ]
            return any(m in body or m in title for m in markers)
        except Exception:
            return False

    def _has_no_flights_text(self, body: str) -> bool:
        text = body.lower()
        patterns = [
            # ── Webjet-specific no-flights messages (client-side SPA render) ──
            # Seen in debug: "Hmm, that route's a tricky one"
            r"that route.*tricky",
            r"hmm.*tricky",
            # Seen in debug: "We couldn't find any flights for your search"
            r"couldn't find any flights",
            r"could not find any flights",
            # Seen in debug: "Maybe the dates need a shuffle"
            r"dates need a shuffle",
            # Seen in debug: "airlines simply aren't flying that route"
            r"aren't flying that route",
            r"are not flying that route",
            # ── Generic no-flights patterns ────────────────────────────────────
            r"no flights? available",
            r"no fares? available",
            r"no results? found",
            r"unable to find flights?",
            r"no flights? found",
            r"flights? not available",
            r"sorry.*no.*flights?",
            r"there are no flights",
            r"0 flights found",
        ]
        return any(re.search(p, text) for p in patterns)

    async def _page_has_error(self, page) -> bool:
        try:
            body = (await page.inner_text("body", timeout=3000)).lower()
            return any(m in body for m in [
                "500 internal server error",
                "404 not found",
                "service unavailable",
                "bad gateway",
            ])
        except Exception:
            return False

    # ── Navigation and page-load wait ───────────────────────

    # ── Webjet error-page detection ──────────────────────────

    def _is_webjet_error_page(self, current_url: str) -> bool:
        """
        When no flights exist for a date, Webjet redirects to:
          https://services.webjet.com.au/web/content/error
        The page title is "Error" and h1 says "Bit of a hiccup, but we're on it".
        This is a definitive NO_FARE signal — no retries needed.
        """
        return "/web/content/error" in current_url

    async def navigate_and_wait(self, page, url: str,
                                timeout: int = PAGE_LOAD_TIMEOUT) -> str:
        """
        Navigate to the Webjet URL and wait for the page to render flights.

        Returns one of:
          "loaded"     — matrix table with result rows present
          "no_flights" — Webjet error redirect  OR  empty matrix table
          "blocked"    — Cloudflare/captcha not cleared within timeout
          "error"      — generic server error
          "timeout"    — nothing useful appeared within timeout seconds
        """
        print("   🌐 Navigating to Webjet URL...")
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=70000)
        except Exception as exc:
            exc_str = str(exc)
            print(f"   ⚠️  Initial navigation exception: {exc}")
            # Bright Data peer-unavailable error: the proxy pool has no routing
            # nodes free at this moment.  The page will be blank — bail out
            # immediately for a fast retry rather than spinning for 90 s.
            if any(marker in exc_str for marker in (
                "no_peer", "no_peers", "a2a_exception", "a2a_tun_open",
            )):
                print("   ⚠️  Bright Data peer unavailable (no_peers) — aborting")
                return "error"
            # Other navigation errors (e.g. transient network blip):
            # continue — Bright Data may still render the page.

        # ── IMMEDIATE CHECK: Webjet error-page redirect ───────
        # When no flights exist for a date, Webjet redirects instantly to
        # /web/content/error ("Bit of a hiccup, but we're on it").
        # Detect this right after domcontentloaded — no need to wait.
        # Marked "suspect" because the same redirect can occur when Webjet's
        # anti-bot system blocks the request; a fresh IP should verify.
        await asyncio.sleep(1)  # allow redirect to settle
        try:
            if self._is_webjet_error_page(page.url):
                print(
                    "   ℹ️  Webjet redirected to error page "
                    f"({page.url}) — no flights for this date"
                )
                return "no_flights_suspect"
        except Exception:
            pass

        loop = asyncio.get_running_loop()
        deadline = loop.time() + timeout
        captcha_attempted  = False
        check_interval     = 1.5   # seconds between polls

        # ── Debounce for no-flights signals ───────────────────────────────────
        # Webjet is a React SPA. During page load it transiently shows error
        # states (the "hiccup" error body, the "no flights" SPA text, or an
        # empty tbody) even on dates that DO have flights.  A single-snapshot
        # decision causes false no-fare detections depending on IP speed.
        #
        # Fix: require every no-flights signal to persist for _DEBOUNCE_SECS
        # across consecutive polls before we trust it.  If tr.result-row ever
        # appears during that window, success wins immediately — the debounce
        # only affects dates that are genuinely flight-less (signal stays stable
        # for the full duration).
        #
        # Typical timings observed:
        #   Flights load         : 2–8 s  →  caught by immediate success check
        #   Genuine no-flights   : signal persists indefinitely → caught after debounce
        #   Transient error state: clears within 1–4 s → debounce resets, rows load
        _DEBOUNCE_SECS: float     = 6.0
        _no_flights_since: float | None = None   # when signal was first seen
        _no_flights_reason: str   = ""

        while loop.time() < deadline:

            # ── 1. SUCCESS — immediate, highest priority ──────────────────────
            # If flight rows are present, return right away regardless of any
            # no-flights signal that may have been accumulating.
            try:
                result_rows = await page.query_selector_all("tr.result-row")
                if result_rows:
                    print(
                        f"   ✅ Matrix table loaded "
                        f"({len(result_rows)} flight row(s))"
                    )
                    return "loaded"
            except Exception:
                pass

            # ── 2. Collect no-flights evidence for this poll ──────────────────
            nf_reason_this_poll: str = ""

            # (a) URL redirect to Webjet error page
            try:
                if self._is_webjet_error_page(page.url):
                    nf_reason_this_poll = "Webjet error-page URL"
            except Exception:
                pass

            # (b) Matrix table rendered but tbody has no result rows
            if not nf_reason_this_poll:
                try:
                    matrix = await page.query_selector("table.matrix-table")
                    if matrix:
                        tbody = await matrix.query_selector("tbody.body")
                        if tbody:
                            tbody_rows = await tbody.query_selector_all(
                                "tr.result-row"
                            )
                            if not tbody_rows:
                                nf_reason_this_poll = "matrix table empty tbody"
                except Exception:
                    pass

            # (c) Body text — Webjet-specific and generic no-flights messages
            if not nf_reason_this_poll:
                try:
                    body = await page.inner_text("body", timeout=2000)
                    if self._has_no_flights_text(body):
                        nf_reason_this_poll = "no-flights body text"
                    elif "bit of a hiccup" in body.lower():
                        nf_reason_this_poll = "hiccup error body text"
                except Exception:
                    pass

            # ── 3. Debounce: accumulate or reset the signal ───────────────────
            if nf_reason_this_poll:
                if _no_flights_since is None:
                    # First poll to show this signal — start the clock
                    _no_flights_since  = loop.time()
                    _no_flights_reason = nf_reason_this_poll
                    print(
                        f"   ⏳ No-flights signal ({nf_reason_this_poll}) — "
                        f"confirming for {_DEBOUNCE_SECS:.0f}s..."
                    )
                else:
                    elapsed = loop.time() - _no_flights_since
                    if elapsed >= _DEBOUNCE_SECS:
                        # Signal has been present for the full debounce window
                        # without flight rows ever appearing → genuine no-flights.
                        # "no-flights body text" = SPA confirmed → certain.
                        # "hiccup error body text" / URL redirect = ambiguous
                        # (could be Webjet anti-bot) → suspect, needs verification.
                        print(
                            f"   ℹ️  No flights confirmed "
                            f"({_no_flights_reason} held {elapsed:.1f}s)"
                        )
                        if _no_flights_reason == "no-flights body text":
                            return "no_flights"          # certain — SPA said so
                        return "no_flights_suspect"      # ambiguous — verify with fresh IP
            else:
                # No no-flights signal this poll — reset debounce
                if _no_flights_since is not None:
                    print(
                        "   ↩️  No-flights signal cleared — "
                        "SPA may still be loading flights"
                    )
                    _no_flights_since  = None
                    _no_flights_reason = ""

            # ── 4. Server error ───────────────────────────────────────────────
            if await self._page_has_error(page):
                print("   ⚠️  Server error page detected")
                return "error"

            # ── 5. Cloudflare / captcha challenge ─────────────────────────────
            if await self._page_is_cloudflare(page):
                if not captcha_attempted:
                    print(
                        "   🛡️  Cloudflare challenge detected — "
                        "Bright Data solving..."
                    )
                remaining_ms = max(5000, int((deadline - loop.time()) * 1000))
                cap_status = await self.wait_for_brightdata_captcha(
                    page, detect_timeout=min(30000, remaining_ms)
                )
                captcha_attempted = True
                if cap_status == "failed":
                    # Bright Data hard-failed — try CapSolver as a fallback.
                    print(
                        "   🔄 Bright Data challenge hard-failed — "
                        "trying CapSolver as fallback..."
                    )
                    remaining_seconds = max(1.0, deadline - loop.time())
                    if not await self._try_capsolver_webjet(page, remaining_seconds):
                        print("   ❌ Cloudflare challenge failed and CapSolver fallback also failed")
                        return "blocked"
                    # CapSolver injected a token — give page time to reload
                    await asyncio.sleep(5)
                # Reset debounce after captcha — page state will change
                _no_flights_since  = None
                _no_flights_reason = ""
                await asyncio.sleep(2)
                continue

            await asyncio.sleep(check_interval)

        print(f"   ⚠️  Page load timed out after {timeout}s")
        return "timeout"

    # ── Flight extraction ────────────────────────────────────

    async def extract_flights(self, page, date_str: str,
                              origin: str, dest: str) -> list[dict]:
        """
        Extract flights from the rendered Webjet page.

        Layer 1: table.matrix-table  (primary — uses known Webjet HTML structure)
        Layer 2: tr.result-row scan  (fallback if table reference is lost)
        Layer 3: body text scan      (last resort)
        """
        now     = wj_now()
        ck_date = now.strftime("%d-%m-%Y")
        ck_time = now.strftime("%H:%M:%S")

        # ── Layer 1: matrix table ─────────────────────────────
        try:
            table = await page.query_selector("table.matrix-table")
            if table:
                data = await self._extract_matrix_table(
                    table, date_str, origin, dest, ck_date, ck_time
                )
                if data:
                    print(f"   ✅ Matrix table extraction: {len(data)} row(s)")
                    return data
        except Exception as exc:
            print(f"   ⚠️  Matrix table extraction error: {exc}")

        # ── Layer 2: result-row fallback ──────────────────────
        try:
            rows = await page.query_selector_all("tr.result-row")
            if rows:
                print(f"   🔍 result-row fallback: {len(rows)} row(s)")
                data = await self._extract_result_rows(
                    page, rows, date_str, origin, dest, ck_date, ck_time
                )
                if data:
                    print(f"   ✅ result-row extraction: {len(data)} row(s)")
                    return data
        except Exception as exc:
            print(f"   ⚠️  result-row extraction error: {exc}")

        # ── Layer 3: body text scan ───────────────────────────
        print("   ⚠️  DOM extraction failed — body text scan fallback")
        try:
            body = await page.inner_text("body", timeout=5000)
        except Exception as exc:
            print(f"   ⚠️  Body text scan failed: {exc}")
            return []

        data: list[dict] = []
        seen: set[str]   = set()
        # Find prices near times (Webjet body text has "4:00 pm … $775" patterns)
        # Collect all times and prices in document order
        times  = re.findall(r"\d{1,2}:\d{2}\s*[aApP][mM]", body)
        prices = re.findall(r"\$\s*\d+(?:[.,]\d+)?", body)
        fare_class = extract_fare_class_from_text(body)

        for i, t in enumerate(times):
            price = prices[i] if i < len(prices) else "N/A"
            key = f"ZL-{t}"
            if key not in seen:
                seen.add(key)
                data.append(self._row(
                    ck_date, ck_time, "ZL", date_str,
                    t, origin, dest, price, fare_class
                ))

        if data:
            print(f"   ✅ Body scan fallback: {len(data)} row(s)")
        return data

    async def _extract_matrix_table(
        self,
        table,
        date_str: str,
        origin: str,
        dest: str,
        ck_date: str,
        ck_time: str,
    ) -> list[dict]:
        """
        Parse the Webjet matrix table:

          thead
            th.header-col  (one per fare class)
              img[data-testid="flight-carrier-ZL-Flex"]  → airline code
              span.fare-type-text                         → "Flex"
          tbody
            tr.result-row  (one per departure time)
              td.left-header
                span.time-depart span  → "4:00 pm"
                span.time-arrive span  → "7:15 pm"
                span.time-duration     → "3h 15m"
                span.stops             → "1"
              td.fare-price  (one per fare class column)
                span.price-text        → "$775"
        """
        data: list[dict] = []
        seen: set[str]   = set()

        # ── Read fare-class columns from <thead> ──────────────
        header_cols = await table.query_selector_all("thead th.header-col")
        fare_columns: list[tuple[str, str]] = []  # (airline_code, fare_type)

        for col in header_cols:
            try:
                # Airline code from img data-testid
                # e.g. "flight-carrier-ZL-Flex" → "ZL"
                airline_code = "ZL"
                img = await col.query_selector("img[data-testid]")
                if img:
                    testid = (await img.get_attribute("data-testid")) or ""
                    parts = testid.split("-")
                    # ["flight", "carrier", "ZL", "Flex"]
                    if (len(parts) >= 3
                            and parts[0] == "flight"
                            and parts[1] == "carrier"):
                        airline_code = parts[2]

                # Fare type from span.fare-type-text
                fare_el = await col.query_selector("span.fare-type-text")
                fare_type = (
                    (await fare_el.inner_text()).strip()
                    if fare_el else "Economy"
                )
                fare_columns.append((airline_code, fare_type))
            except Exception:
                fare_columns.append(("ZL", "Economy"))

        print(f"   📋 Fare columns: {fare_columns}")
        if not fare_columns:
            return data

        # ── Read flight rows from <tbody> ─────────────────────
        result_rows = await table.query_selector_all("tbody tr.result-row")
        print(f"   🔍 Matrix result rows: {len(result_rows)}")

        for row in result_rows:
            try:
                # Departure time
                dep_el = await row.query_selector(
                    "td.left-header span.time-depart span"
                )
                dep_time = (
                    (await dep_el.inner_text()).strip() if dep_el else "-"
                )

                # Arrival time (logged for context, not stored as its own column)
                arr_el = await row.query_selector(
                    "td.left-header span.time-arrive span"
                )
                arr_time = (
                    (await arr_el.inner_text()).strip() if arr_el else ""
                )

                # Stops
                stops_el = await row.query_selector(
                    "td.left-header span.stops"
                )
                stops = (
                    (await stops_el.inner_text()).strip() if stops_el else ""
                )

                # Price cells — one td.fare-price per fare-class column
                price_cells = await row.query_selector_all("td.fare-price")

                for col_idx, cell in enumerate(price_cells):
                    try:
                        price_el = await cell.query_selector("span.price-text")
                        if not price_el:
                            continue
                        price_raw = (await price_el.inner_text()).strip()
                        if not price_raw or price_raw in ("-", "", "—"):
                            continue
                        price = (
                            price_raw
                            if price_raw.startswith("$")
                            else f"${price_raw}"
                        )

                        if col_idx < len(fare_columns):
                            airline_code, fare_type = fare_columns[col_idx]
                        else:
                            airline_code, fare_type = "ZL", "Economy"

                        key = f"{airline_code}-{dep_time}-{fare_type}"
                        if key in seen:
                            continue
                        seen.add(key)

                        stops_note = f" ({stops} stop{'s' if stops not in ('0','1') else ''})" if stops else ""
                        print(
                            f"      ✈️  {airline_code}  dep={dep_time}"
                            f"  arr={arr_time}{stops_note}"
                            f"  price={price}  class={fare_type}"
                        )
                        data.append(self._row(
                            ck_date, ck_time,
                            airline_code,
                            date_str,
                            dep_time,
                            origin, dest,
                            price,
                            fare_type,
                        ))
                    except Exception:
                        continue
            except Exception as exc:
                print(f"   ⚠️  Row parse error: {exc}")
                continue

        return data

    async def _extract_result_rows(
        self,
        page,
        rows,
        date_str: str,
        origin: str,
        dest: str,
        ck_date: str,
        ck_time: str,
    ) -> list[dict]:
        """
        Fallback extractor: parse tr.result-row elements directly
        without a full table reference (e.g. if matrix-table selector fails).
        """
        data: list[dict] = []
        seen: set[str]   = set()

        # Try to read fare-type labels from page-level elements
        fare_labels: list[str] = []
        try:
            fare_els = await page.query_selector_all("span.fare-type-text")
            fare_labels = [
                (await el.inner_text()).strip() for el in fare_els
            ]
        except Exception:
            pass
        if not fare_labels:
            fare_labels = ["Economy"]

        for row in rows:
            try:
                dep_el = await row.query_selector(
                    "span.time-depart span"
                )
                dep_time = (
                    (await dep_el.inner_text()).strip() if dep_el else "-"
                )

                price_cells = await row.query_selector_all("td.fare-price")
                for col_idx, cell in enumerate(price_cells):
                    try:
                        price_el = await cell.query_selector("span.price-text")
                        if not price_el:
                            continue
                        price_raw = (await price_el.inner_text()).strip()
                        if not price_raw or price_raw in ("-", ""):
                            continue
                        price = (
                            price_raw
                            if price_raw.startswith("$")
                            else f"${price_raw}"
                        )
                        fare_type = (
                            fare_labels[col_idx]
                            if col_idx < len(fare_labels)
                            else "Economy"
                        )
                        key = f"ZL-{dep_time}-{fare_type}"
                        if key not in seen:
                            seen.add(key)
                            data.append(self._row(
                                ck_date, ck_time, "ZL", date_str,
                                dep_time, origin, dest, price, fare_type,
                            ))
                    except Exception:
                        continue
            except Exception:
                continue

        return data

    # ── Row helpers ──────────────────────────────────────────

    def _row(self, ck_date, ck_time, airline, dep_date, dep_time,
             orig, dest, price, fare_class="Economy",
             source="Webjet", status=STATUS_SUCCESS, comment="",
             retry_count=0, debug_artifacts="") -> dict:
        return {
            "Date Checked":      ck_date,
            "Time Checked":      ck_time,
            "Airline":           airline,
            "Date of Departure": dep_date,
            "Time of Departure": dep_time,
            "Origin":            orig,
            "Destination":       dest,
            "Fare Price":        price,
            "Fare Class":        fare_class,
            "Source":            source,
            "Run ID":            RUN_ID,
            "Status":            status,
            "Comment":           comment,
            "Retry Count":       retry_count,
            "Debug Artifacts":   debug_artifacts,
        }

    def _no_flight_row(self, date_str, orig, dest,
                       comment="No flights available on Webjet for this date") -> dict:
        now = wj_now()
        return self._row(
            now.strftime("%d-%m-%Y"), now.strftime("%H:%M:%S"),
            "no flight", date_str, "-", orig, dest, "N/A",
            fare_class="", source="Webjet",
            status=STATUS_NO_FARE, comment=comment,
        )

    def _failed_row(self, date_str, orig, dest, status, comment,
                    retry_count=0, debug_artifacts="") -> dict:
        now = wj_now()
        return self._row(
            now.strftime("%d-%m-%Y"), now.strftime("%H:%M:%S"),
            "scrape failed", date_str, "-", orig, dest, "N/A",
            fare_class="", source="Webjet - failed",
            status=status, comment=comment,
            retry_count=retry_count, debug_artifacts=debug_artifacts,
        )

    # ── Retry backoff ────────────────────────────────────────

    def _backoff_delay(self, attempt: int) -> float:
        if self.retry_backoff <= 0:
            return 0.0
        jitter = random.uniform(0.0, min(3.0, self.retry_backoff))
        return min(
            self.max_retry_backoff,
            self.retry_backoff * (2 ** max(0, attempt - 1)) + jitter
        )

    # ── Result writer ────────────────────────────────────────

    def _write_result(self, origin: str, dest: str,
                      date_str: str, result: JobResult):
        rows = result.rows
        if not rows:
            rows = [self._failed_row(
                date_str, origin, dest,
                result.status,
                result.comment or "Job failed without parsed rows",
                result.retry_count,
                " | ".join(result.debug_artifacts),
            )]
        if OUTPUT_STORE:
            OUTPUT_STORE.write_job_rows(origin, dest, date_str, rows)
        else:
            _append_rows(rows)

    # ── Single date scrape ───────────────────────────────────

    async def scrape_date_once(self, page, origin: str, dest: str,
                               target_dt: datetime, attempt: int) -> JobResult:
        """
        Scrape a single date using an already-open *page*.
        The caller owns the page lifecycle — no context is created or closed here.
        Reusing the same page across dates allows Chromium's in-memory HTTP cache
        to serve static React SPA assets (JS bundles, CSS) after the first load,
        cutting per-date bandwidth by ~80%.
        """
        date_str = output_date(target_dt)
        url      = build_webjet_url(origin, dest, target_dt)
        label    = f"{RUN_ID}_{origin}_{dest}_{date_str}_attempt_{attempt}"

        load_result = await self.navigate_and_wait(
            page, url, timeout=PAGE_LOAD_TIMEOUT
        )

        # ── Blocked ────────────────────────────────────────────
        if load_result == "blocked":
            artifacts = await self.save_debug(
                page, f"{label}_blocked",
                {"origin": origin, "dest": dest,
                 "date": date_str, "url": url},
            )
            return JobResult(
                STATUS_BLOCKED, [],
                "Cloudflare/captcha challenge persisted",
                retryable=True, retry_count=attempt - 1,
                debug_artifacts=artifacts,
            )

        # ── Server error ───────────────────────────────────────
        if load_result == "error":
            artifacts = await self.save_debug(
                page, f"{label}_server_error",
                {"origin": origin, "dest": dest,
                 "date": date_str, "url": url},
            )
            return JobResult(
                STATUS_SITE_UNAVAILABLE, [],
                "Webjet returned a server error page",
                retryable=True, retry_count=attempt - 1,
                debug_artifacts=artifacts,
            )

        # ── No flights (certain — SPA body text confirmed) ─────
        if load_result == "no_flights":
            row = self._no_flight_row(
                date_str, origin, dest,
                "Webjet page reports no flights available for this date",
            )
            return JobResult(
                STATUS_NO_FARE, [row],
                "No flights available", retryable=False,
                retry_count=attempt - 1,
                needs_ip_verification=False,
            )

        # ── No flights (suspect — URL redirect or hiccup) ──────
        # Could be Webjet anti-bot blocking rather than genuine no flights.
        # Caller should reconnect for a fresh IP and verify.
        if load_result == "no_flights_suspect":
            row = self._no_flight_row(
                date_str, origin, dest,
                "Webjet page reports no flights available for this date",
            )
            return JobResult(
                STATUS_NO_FARE, [row],
                "No flights available", retryable=False,
                retry_count=attempt - 1,
                needs_ip_verification=True,
            )

        # ── Timeout ────────────────────────────────────────────
        if load_result == "timeout":
            artifacts = await self.save_debug(
                page, f"{label}_timeout",
                {"origin": origin, "dest": dest,
                 "date": date_str, "url": url},
            )
            return JobResult(
                STATUS_TIMEOUT, [],
                f"Page load timed out after {PAGE_LOAD_TIMEOUT}s",
                retryable=True, retry_count=attempt - 1,
                debug_artifacts=artifacts,
            )

        # ── Page loaded — extract flights ──────────────────────
        flights = await self.extract_flights(page, date_str, origin, dest)
        real = [f for f in flights if f.get("Airline") != "no flight"]

        if real:
            comment = "Flight data parsed from Webjet flight matrix page"
            for f in real:
                f["Status"]      = STATUS_SUCCESS
                f["Comment"]     = comment
                f["Retry Count"] = attempt - 1
            return JobResult(
                STATUS_SUCCESS, real, comment,
                retryable=False, retry_count=attempt - 1,
            )

        # Table was loaded but extraction returned nothing.
        # A rendered-but-empty tbody means the route has no flights.
        # Any other case is a structure/parse issue.
        try:
            matrix = await page.query_selector("table.matrix-table")
            if matrix:
                tbody = await matrix.query_selector("tbody.body")
                rows_in_table = (
                    await matrix.query_selector_all("tr.result-row")
                ) if tbody else []
                if tbody and len(rows_in_table) == 0:
                    row = self._no_flight_row(date_str, origin, dest,
                        "Matrix table present but no result rows (no flights)")
                    return JobResult(
                        STATUS_NO_FARE, [row],
                        "Matrix table has no flight rows for this date",
                        retryable=False, retry_count=attempt - 1,
                    )
        except Exception:
            pass

        # Generic no-flights text check
        try:
            body = await page.inner_text("body", timeout=3000)
            if self._has_no_flights_text(body):
                row = self._no_flight_row(date_str, origin, dest)
                return JobResult(
                    STATUS_NO_FARE, [row],
                    "No flights found (post-extract check)",
                    retryable=False, retry_count=attempt - 1,
                )
        except Exception:
            pass

        artifacts = await self.save_debug(
            page, f"{label}_no_extract",
            {"origin": origin, "dest": dest, "date": date_str,
             "url": url,
             "reason": "Matrix table loaded but flight extraction returned no rows"},
        )
        return JobResult(
            STATUS_STRUCTURE, [],
            "Webjet matrix page loaded but flight extraction returned no rows",
            retryable=True, retry_count=attempt - 1,
            debug_artifacts=artifacts,
        )

    # ── Scrape with retries ──────────────────────────────────

    async def scrape_date_with_retries(self, browser, origin: str,
                                       dest: str, target_dt: datetime,
                                       effective_browser_ref: list | None = None,
                                       effective_ctx_ref:     list | None = None,
                                       effective_page_ref:    list | None = None,
                                       ) -> JobResult:
        """
        Try scraping a single date up to max_attempts times.
        Reconnects to Bright Data for a fresh IP on BLOCKED/SITE_UNAVAILABLE
        or on a suspect NO_FARE when the IP has not yet proved itself clean.

        effective_browser_ref / effective_ctx_ref / effective_page_ref are
        mutable [value] lists so callers see any reconnection that happens here.
        """
        date_str     = output_date(target_dt)
        last_result: JobResult | None = None

        eff      = effective_browser_ref if effective_browser_ref is not None else [browser]
        eff_ctx  = effective_ctx_ref   # may be None when caller doesn't use shared ctx
        eff_page = effective_page_ref  # may be None when caller doesn't use shared page

        for attempt in range(1, self.max_attempts + 1):
            print(
                f"   🔁 Attempt {attempt}/{self.max_attempts}: "
                f"{origin}→{dest} {date_str}"
            )
            try:
                result = await asyncio.wait_for(
                    self.scrape_date_once(
                        eff_page[0] if eff_page else eff[0],
                        origin, dest, target_dt, attempt,
                    ),
                    timeout=self.job_timeout,
                )
                result.retry_count = attempt - 1
                last_result = result

                # ── SUCCESS → mark IP clean, done immediately ───────────
                if result.status == STATUS_SUCCESS:
                    self._ip_clean = True  # this IP can reach Webjet fine
                    return result

                # ── NO_FARE (suspect signal) → verify with fresh IP ────
                # Webjet's anti-bot system shows persistent error-page URL
                # redirects and "hiccup" messages for blocked IPs — identical
                # to a genuine "no flights" response.  Only reconnect when:
                #   • the signal is ambiguous (needs_ip_verification), AND
                #   • the IP has NOT already proved itself clean this session.
                # Once the IP has loaded ≥1 SUCCESS we know it isn't blocked —
                # subsequent URL redirects / hiccup errors are genuine Webjet
                # "no flights" responses and need no second opinion.
                if (result.status == STATUS_NO_FARE
                        and result.needs_ip_verification
                        and not self._ip_clean
                        and attempt < self.max_attempts):
                    print(
                        f"   🔍 No-flights (suspect) on attempt {attempt} — "
                        "reconnecting for IP-verification retry..."
                    )
                    ok = await self._reconnect_with_new_page(
                        eff, eff_ctx, eff_page, "verify"
                    )
                    if ok:
                        print("   ✅ Fresh Bright Data IP — verifying no-flights...")
                    delay = self._backoff_delay(attempt)
                    if delay > 0:
                        print(f"   ⏳ Backoff {delay:.1f}s before verification attempt")
                        await asyncio.sleep(delay)
                    continue  # → next attempt with fresh IP + page

                # ── All other completed statuses (NO_FARE confirmed after
                #    all attempts, SITE_UNAVAILABLE, etc.) ─────────────────
                if result.completed:
                    return result
                if not result.retryable:
                    return result
                print(f"   ⚠️  {result.status}: {result.comment}")

            except asyncio.TimeoutError:
                comment = f"Hard timeout ({self.job_timeout}s) exceeded"
                print(f"   ⏱️  {comment}")
                last_result = JobResult(
                    STATUS_TIMEOUT, [], comment,
                    retryable=True, retry_count=attempt - 1,
                )

            except Exception as exc:
                comment = f"Unhandled exception: {exc}"
                print(f"   ❌ {comment}")
                traceback.print_exc()
                last_result = JobResult(
                    STATUS_FAILED, [], comment,
                    retryable=True, retry_count=attempt - 1,
                )

            if attempt < self.max_attempts:
                # Reconnect for a fresh IP / routing path on:
                #   BLOCKED          — Cloudflare/captcha; need a different IP
                #   SITE_UNAVAILABLE — includes Bright Data "no_peers" errors;
                #                      a fresh session routes through different
                #                      peer nodes and usually clears the issue
                if last_result and last_result.status in (
                    STATUS_BLOCKED, STATUS_SITE_UNAVAILABLE
                ):
                    reason = (
                        "blocked" if last_result.status == STATUS_BLOCKED
                        else "site/peer unavailable"
                    )
                    print(
                        f"   🔄 Attempt {attempt} {reason} — "
                        "reconnecting Bright Data for fresh session..."
                    )
                    ok = await self._reconnect_with_new_page(
                        eff, eff_ctx, eff_page, reason
                    )
                    if ok:
                        print("   ✅ Fresh Bright Data session for next attempt")

                delay = self._backoff_delay(attempt)
                if delay > 0:
                    print(f"   ⏳ Backoff {delay:.1f}s before next attempt")
                    await asyncio.sleep(delay)

        comment = f"Job failed after {self.max_attempts} attempt(s)"
        if last_result:
            comment = f"{comment}: {last_result.comment}"
        return JobResult(
            last_result.status if last_result else STATUS_FAILED,
            last_result.rows if last_result else [],
            comment,
            retryable=True,
            retry_count=max(0, self.max_attempts - 1),
            debug_artifacts=last_result.debug_artifacts if last_result else [],
        )

    # ── Cleanup retry pass ───────────────────────────────────

    async def _run_cleanup_retry_pass(self, browser, origin: str,
                                      dest: str, dates: list[datetime]):
        """
        After the main run completes, retry any dates still showing
        retryable failures using a freshly connected Bright Data session.
        """
        if not OUTPUT_STORE:
            return

        date_strs   = [output_date(dt) for dt in dates]
        failed_strs = OUTPUT_STORE.failed_route_dates(origin, dest, date_strs)
        if not failed_strs:
            print(
                f"\n   ✅ Cleanup pass: all {len(dates)} dates completed "
                f"for {origin}→{dest}"
            )
            return

        failed_set   = set(failed_strs)
        failed_dates = [dt for dt in dates if output_date(dt) in failed_set]

        print(f"\n{'─'*60}")
        print(
            f"🔄 Cleanup retry pass: {len(failed_dates)} date(s) still "
            f"failed for {origin}→{dest}"
        )
        print(
            f"   Dates: "
            f"{', '.join(failed_strs[:10])}{'...' if len(failed_strs) > 10 else ''}"
        )
        print(f"{'─'*60}")

        print("   🔌 Reconnecting to Bright Data for fresh IP (cleanup)...")
        await asyncio.sleep(5)
        cleanup_browser = await self._reconnect_browser()
        cleanup_b = cleanup_browser if cleanup_browser else browser

        eff      = [cleanup_b]
        eff_ctx  = [None]
        eff_page = [None]

        succeeded = still_failed = 0
        try:
            ctx, page = await self._new_context(cleanup_b)
            eff_ctx[0]  = ctx
            eff_page[0] = page

            for target_dt in failed_dates:
                date_str = output_date(target_dt)
                if OUTPUT_STORE.job_completed(origin, dest, date_str):
                    print(f"   ↩️  {date_str}: completed since scan — skipping.")
                    continue

                print(f"\n   🔁 Cleanup retry: {origin}→{dest} {date_str}")
                try:
                    result = await self.scrape_date_with_retries(
                        eff[0], origin, dest, target_dt,
                        effective_browser_ref=eff,
                        effective_ctx_ref=eff_ctx,
                        effective_page_ref=eff_page,
                    )
                except Exception as exc:
                    result = JobResult(
                        STATUS_FAILED, [],
                        f"Cleanup retry exception: {exc}",
                        retryable=True,
                    )

                self._write_result(origin, dest, date_str, result)
                if result.completed:
                    print(f"   ✅ Cleanup retry succeeded: {date_str} ({result.status})")
                    succeeded += 1
                else:
                    print(
                        f"   ❌ Cleanup retry still failed: {date_str} — "
                        f"{result.status}: {result.comment}"
                    )
                    still_failed += 1
                    if result.status in (STATUS_BLOCKED, STATUS_SITE_UNAVAILABLE):
                        ok = await self._reconnect_with_new_page(
                            eff, eff_ctx, eff_page, "cleanup"
                        )
                        if not ok:
                            print(
                                "   ❌ Cleanup reconnect failed — "
                                "stopping cleanup pass."
                            )
                            break

        finally:
            if eff_ctx[0]:
                await self._close_context(eff_ctx[0], "cleanup context")

        print(f"\n   📋 Cleanup pass done: {succeeded} recovered, "
              f"{still_failed} still failed.")

    # ── Route runner ─────────────────────────────────────────

    async def run_route(self, origin: str, dest: str):
        origin_name = AIRPORT_MAP.get(origin, origin)
        dest_name   = AIRPORT_MAP.get(dest, dest)
        dates       = build_date_list()

        print(f"\n{'█'*60}")
        print(f"  ROUTE  : {origin} ({origin_name}) → {dest} ({dest_name})")
        print(f"  Window : {output_date(dates[0])} → {output_date(dates[-1])}")
        print(f"  Dates  : {len(dates)}")
        print(f"  Output : {OUTPUT_EXCEL}")
        print(f"  Retry  : attempts={self.max_attempts}, "
              f"final_rounds={self.final_retry_rounds}, "
              f"job_timeout={self.job_timeout}s")
        print(f"{'█'*60}")

        async with async_playwright() as p:
            self._pw_instance = p
            print("🔌 Connecting to Bright Data Browser API...")
            browser = None
            try:
                browser = await p.chromium.connect_over_cdp(BD_BROWSER_WSS)
                print("✅ Bright Data Browser API connected.")
            except Exception as exc:
                print(f"❌ Could not connect to Bright Data: {exc}")
                skipped = 0
                for dt in dates:
                    date_str = output_date(dt)
                    # NEVER overwrite dates that already have good data.
                    # A connection failure must not destroy previously scraped rows.
                    if OUTPUT_STORE and OUTPUT_STORE.job_completed(
                            origin, dest, date_str):
                        skipped += 1
                        continue
                    self._write_result(
                        origin, dest, date_str,
                        JobResult(
                            STATUS_SITE_UNAVAILABLE, [],
                            f"Cannot connect to Bright Data Browser API: {exc}",
                            retryable=True,
                        ),
                    )
                if skipped:
                    print(
                        f"   ℹ️  {skipped} already-completed date(s) preserved "
                        f"(not overwritten by connection failure)."
                    )
                return

            failed_jobs: list[datetime] = []
            eff      = [browser]  # mutable so reconnections persist across calls
            eff_ctx  = [None]     # shared browser context (HTTP cache persists)
            eff_page = [None]     # shared page within that context

            self._ip_clean = False  # reset at start of each route

            try:
                # One shared context+page for the whole route.
                # Chromium's in-memory cache keeps React SPA assets (JS, CSS)
                # after the first navigation — subsequent dates only transfer
                # dynamic flight data (~0.3 MB) instead of the full ~3 MB.
                ctx, page = await self._new_context(browser)
                eff_ctx[0]  = ctx
                eff_page[0] = page

                # ── Main date loop ─────────────────────────────────
                for idx, target_dt in enumerate(dates, 1):
                    date_str = output_date(target_dt)
                    print(f"\n{'═'*60}")
                    print(f"📅 [{idx}/{len(dates)}]  "
                          f"{target_dt.strftime('%A, %d %b %Y')}")
                    print(f"{'─'*60}")

                    if (self.resume and OUTPUT_STORE
                            and OUTPUT_STORE.job_completed(origin, dest, date_str)):
                        print(f"   ↩️  Resume: already completed — skipping.")
                        continue

                    result = await self.scrape_date_with_retries(
                        eff[0], origin, dest, target_dt,
                        effective_browser_ref=eff,
                        effective_ctx_ref=eff_ctx,
                        effective_page_ref=eff_page,
                    )
                    self._write_result(origin, dest, date_str, result)

                    if result.completed:
                        if result.status == STATUS_SUCCESS:
                            print(f"   ✅ {len(result.rows)} flight row(s) saved.")
                        else:
                            print(f"   ℹ️  No fare: {result.comment}")
                    else:
                        print(f"   ❌ {result.status} — {result.comment}")
                        failed_jobs.append(target_dt)

                # ── Final retry rounds ─────────────────────────────
                for retry_round in range(1, self.final_retry_rounds + 1):
                    retry_targets = [
                        dt for dt in failed_jobs
                        if not (OUTPUT_STORE and OUTPUT_STORE.job_completed(
                            origin, dest, output_date(dt)))
                    ]
                    if not retry_targets:
                        break

                    print(f"\n{'═'*60}")
                    print(
                        f"🔁 Final retry round "
                        f"{retry_round}/{self.final_retry_rounds}: "
                        f"{len(retry_targets)} job(s)"
                    )
                    print(f"{'═'*60}")
                    still_failed: list[datetime] = []

                    for target_dt in retry_targets:
                        date_str = output_date(target_dt)
                        result = await self.scrape_date_with_retries(
                            eff[0], origin, dest, target_dt,
                            effective_browser_ref=eff,
                            effective_ctx_ref=eff_ctx,
                            effective_page_ref=eff_page,
                        )
                        self._write_result(origin, dest, date_str, result)
                        if result.completed:
                            print(f"   ✅ Retry OK: {date_str}")
                        else:
                            print(
                                f"   ❌ Still failed: {date_str} "
                                f"— {result.status}"
                            )
                            still_failed.append(target_dt)

                    failed_jobs = still_failed

                # ── Completeness guard ─────────────────────────────
                if OUTPUT_STORE:
                    missing = []
                    for target_dt in dates:
                        date_str = output_date(target_dt)
                        if not OUTPUT_STORE.job_has_any_row(
                                origin, dest, date_str):
                            missing.append(target_dt)
                            self._write_result(
                                origin, dest, date_str,
                                JobResult(
                                    STATUS_FAILED, [],
                                    "Completeness guard: no output existed "
                                    "for expected route/date",
                                    retryable=True,
                                    retry_count=max(0, self.max_attempts - 1),
                                ),
                            )
                    if missing:
                        print(
                            f"   🚨 Completeness guard added "
                            f"{len(missing)} missing row(s)."
                        )

            except KeyboardInterrupt:
                print("\n⛔ Interrupted.")
            except Exception as exc:
                print(f"\n❌ Fatal route error: {exc}")
                traceback.print_exc()
                if OUTPUT_STORE:
                    for dt in dates:
                        date_str = output_date(dt)
                        if not OUTPUT_STORE.job_has_any_row(origin, dest, date_str):
                            self._write_result(
                                origin, dest, date_str,
                                JobResult(
                                    STATUS_ROUTE_SETUP_FAILED, [],
                                    f"Route-level failure: {exc}",
                                    retryable=True,
                                ),
                            )
            else:
                # Main loop finished without exception — run cleanup pass
                try:
                    await self._run_cleanup_retry_pass(
                        eff[0], origin, dest, dates
                    )
                except KeyboardInterrupt:
                    print("\n⛔ Cleanup pass interrupted.")
                except Exception as exc:
                    print(f"\n⚠️  Cleanup pass error (non-fatal): {exc}")
            finally:
                if eff_ctx[0]:
                    await self._close_context(eff_ctx[0], "route context")
                if browser:
                    await browser.close()

        print(f"\n  📊 Output saved: {OUTPUT_EXCEL}\n")


# ─────────────────────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(
        description="Webjet.com.au scraper — PER-MJK and MJK-PER via Bright Data"
    )
    parser.add_argument(
        "route_codes", nargs="*",
        help="Route pairs e.g. PER MJK  or  MJK PER  (default: both routes)"
    )
    parser.add_argument(
        "--routes",
        help="Comma-separated routes e.g. PER-MJK,MJK-PER"
    )
    parser.add_argument(
        "--days", type=int, default=TOTAL_DAYS,
        help=f"Number of dates to scrape per route (default: {TOTAL_DAYS})"
    )
    parser.add_argument(
        "--start-offset-days", type=int, default=START_OFFSET_DAYS,
        help="Days after today to start (default: 1 = tomorrow)"
    )
    parser.add_argument(
        "--output", default=OUTPUT_EXCEL,
        help=f"Excel output file (default: {OUTPUT_EXCEL})"
    )
    parser.add_argument(
        "--debug-dir", default=DEBUG_DIR,
        help="Directory for debug screenshots/HTML"
    )
    parser.add_argument(
        "--log-dir", default=LOG_DIR,
        help="Directory for cron-friendly run logs"
    )
    parser.add_argument(
        "--run-id", default=RUN_ID,
        help="Resume key written to output (default: YYYYMMDD in Perth TZ)"
    )
    parser.add_argument(
        "--max-attempts", type=int, default=MAX_ATTEMPTS,
        help=f"Attempts per date before final retry queue (default: {MAX_ATTEMPTS})"
    )
    parser.add_argument(
        "--final-retry-rounds", type=int, default=FINAL_RETRY_ROUNDS,
        help=f"End-of-route retry rounds for failed dates (default: {FINAL_RETRY_ROUNDS})"
    )
    parser.add_argument(
        "--retry-backoff", type=float, default=RETRY_BACKOFF_SECONDS,
        help=f"Base retry backoff seconds (default: {RETRY_BACKOFF_SECONDS})"
    )
    parser.add_argument(
        "--max-retry-backoff", type=float, default=MAX_RETRY_BACKOFF_SECS,
        help=f"Maximum retry backoff seconds (default: {MAX_RETRY_BACKOFF_SECS})"
    )
    parser.add_argument(
        "--job-timeout", type=int, default=JOB_TIMEOUT_SECONDS,
        help=f"Hard timeout per date attempt in seconds (default: {JOB_TIMEOUT_SECONDS})"
    )
    parser.add_argument(
        "--page-load-timeout", type=int, default=PAGE_LOAD_TIMEOUT,
        help=f"Max seconds to wait for Webjet page to show flights (default: {PAGE_LOAD_TIMEOUT})"
    )
    parser.add_argument(
        "--inter-route-delay", type=int, default=INTER_ROUTE_DELAY_SECS,
        metavar="SECS",
        help=f"Pause between routes for Bright Data IP rotation (default: {INTER_ROUTE_DELAY_SECS}s)"
    )
    parser.add_argument(
        "--no-resume", action="store_true",
        help="Re-scrape all dates even if already completed for this run ID"
    )
    parser.add_argument(
        "--list", action="store_true",
        help="List supported routes and exit"
    )
    return parser.parse_args()


def parse_routes(ns) -> list[tuple[str, str]]:
    if ns.routes:
        pairs = []
        for token in ns.routes.split(","):
            token = token.strip()
            if not token:
                continue
            parts = [p for p in re.split(r"[-:>]", token.upper()) if p]
            if len(parts) != 2:
                raise ValueError(f"Bad route format: {token!r}")
            pairs.append((parts[0], parts[1]))
        return pairs

    if ns.route_codes:
        codes = [c.upper() for c in ns.route_codes]
        if len(codes) % 2 != 0:
            raise ValueError("Route codes must be supplied in ORIGIN DEST pairs.")
        return list(zip(codes[0::2], codes[1::2]))

    return list(WEBJET_ROUTES)


if __name__ == "__main__":
    ns = parse_args()

    if ns.list:
        print("\nSupported Webjet routes:")
        for o, d in WEBJET_ROUTES:
            cfg = ROUTE_CONFIG[(o, d)]
            print(f"  {o} → {d}  ({cfg['CityFrom']} → {cfg['CityTo']})")
        print()
        sys.exit(0)

    # ── Route selection ───────────────────────────────────────
    try:
        routes_to_run = parse_routes(ns)
    except ValueError as exc:
        print(f"❌ {exc}")
        sys.exit(1)

    bad = [r for r in routes_to_run if r not in WEBJET_ROUTES]
    if bad:
        print(f"❌ Unsupported route(s): {bad}")
        print(f"   Supported: {WEBJET_ROUTES}")
        sys.exit(1)

    # ── Config apply ─────────────────────────────────────────
    TOTAL_DAYS             = max(1, ns.days)
    START_OFFSET_DAYS      = max(0, ns.start_offset_days)
    OUTPUT_EXCEL           = ns.output
    DEBUG_DIR              = ns.debug_dir
    LOG_DIR                = ns.log_dir
    RUN_ID                 = ns.run_id
    MAX_ATTEMPTS           = max(1, ns.max_attempts)
    FINAL_RETRY_ROUNDS     = max(0, ns.final_retry_rounds)
    RETRY_BACKOFF_SECONDS  = max(0.0, ns.retry_backoff)
    MAX_RETRY_BACKOFF_SECS = max(1.0, ns.max_retry_backoff)
    JOB_TIMEOUT_SECONDS    = max(30, ns.job_timeout)
    PAGE_LOAD_TIMEOUT      = max(20, ns.page_load_timeout)
    INTER_ROUTE_DELAY_SECS = max(0, ns.inter_route_delay)
    RESUME_ENABLED         = not ns.no_resume

    log_fh = configure_run_logging(LOG_DIR, RUN_ID)
    OUTPUT_STORE = OutputStore(OUTPUT_EXCEL, RUN_ID)

    scraper    = WebjetScraper(
        debug_dir         = DEBUG_DIR,
        max_attempts      = MAX_ATTEMPTS,
        final_retry_rounds= FINAL_RETRY_ROUNDS,
        retry_backoff     = RETRY_BACKOFF_SECONDS,
        max_retry_backoff = MAX_RETRY_BACKOFF_SECS,
        job_timeout       = JOB_TIMEOUT_SECONDS,
        resume            = RESUME_ENABLED,
    )
    start_dt = build_date_list()[0]

    # ── Run summary ───────────────────────────────────────────
    print("═" * 60)
    print("  WEBJET SCRAPER — RUN SUMMARY")
    print("═" * 60)
    print(f"  📋 Routes      : {len(routes_to_run)}")
    for o, d in routes_to_run:
        cfg = ROUTE_CONFIG[(o, d)]
        print(f"       {o} → {d}  ({cfg['CityFrom']} → {cfg['CityTo']})")
    print(f"  📆 Dates/route : {TOTAL_DAYS}  (from {output_date(start_dt)})")
    print(f"  🗂  Output      : {OUTPUT_EXCEL}")
    print(
        f"  🔁 Resume      : "
        f"{'Yes' if RESUME_ENABLED else 'No'}  |  Run ID: {RUN_ID}"
    )
    print(f"  💾 Checkpoint  : every {CHECKPOINT_EVERY} entries")
    print(f"  🌐 Bright Data : Scraping Browser API (Cloudflare auto-handled)")
    print("═" * 60 + "\n")

    # ── Route execution ───────────────────────────────────────
    route_status: dict[tuple[str, str], str] = {}

    try:
        for route_idx, (origin, dest) in enumerate(routes_to_run):
            try:
                asyncio.run(scraper.run_route(origin, dest))
                route_status[(origin, dest)] = "ok"
            except KeyboardInterrupt:
                print(f"\n⛔ Stopped at {origin}→{dest}.")
                route_status[(origin, dest)] = "interrupted"
                break
            except Exception as exc:
                route_status[(origin, dest)] = (
                    f"ERROR: {type(exc).__name__}: {exc}"
                )
                print(f"\n❌ Route {origin}→{dest} failed: "
                      f"{type(exc).__name__}: {exc}")

            # Inter-route delay — lets Bright Data rotate to a fresh IP
            if route_idx < len(routes_to_run) - 1 and INTER_ROUTE_DELAY_SECS > 0:
                print(
                    f"\n⏸️  Inter-route cooldown: {INTER_ROUTE_DELAY_SECS}s "
                    "(letting Bright Data IP pool rotate)..."
                )
                time.sleep(INTER_ROUTE_DELAY_SECS)
    finally:
        print("\n" + "═" * 60)
        print("  WEBJET SCRAPER — ROUTE RESULTS")
        print("═" * 60)
        for (o, d) in routes_to_run:
            st = route_status.get((o, d), "not reached")
            if st == "ok":
                print(f"  ✅  {o} → {d}")
            elif st == "interrupted":
                print(f"  ⛔  {o} → {d}  (interrupted)")
            else:
                reason = st.replace("ERROR: ", "")
                print(f"  ❌  {o} → {d}")
                print(f"       {reason}")
        print("═" * 60)
        restore_run_logging(log_fh)
